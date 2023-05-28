# pip3 install openpyxl
# https://openpyxl.readthedocs.io/en/stable/index.html
from openpyxl import Workbook, load_workbook

wb = load_workbook("exemplo.xlsx")
type(wb)

# Retorna todas as abas das planilhas
wb.get_sheet_names()

# Cria um objeto que controla uma planilha
sheet = wb["Sheet1"]
sheet = wb.get_sheet_by_name("Sheet1")
type(sheet)
sheet.title

# Métodos básicos para navegação em céluas
sheet["A3"]
sheet["A3"].value

# Podemos acessar as células também utilizando rows e cols
sheet.cell(row=1, column=2)
sheet.cell(row=1, column=2).value

for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)


for row in sheet.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
        print(cell)


# Agrupamento / desagrupamento de células
sheet.merge_cells("A1:D1")

sheet.merge_cells("A2:D2")
sheet.unmerge_cells("A2:D2")
sheet.merge_cells(start_row=2, start_column=1, end_row=4, end_column=4)
sheet.unmerge_cells(start_row=2, start_column=1, end_row=4, end_column=4)
wb.save("formula.xlsx")


# Adicionando linhas e colunas
sheet.insert_rows(7)
sheet.delete_cols(2, 5)  # Deleta as colunas B:F

# Adicionando imagens
from openpyxl.drawing.image import Image

img = Image("catlogo.png")
sheet.add_image(img, "A1")
wb.save("formula.xlsx")
